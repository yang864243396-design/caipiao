"""Excel 边跑边追加。"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from config import EXCEL_HEADERS, REPORTS_DIR
from models import CaseResult


def new_report_path() -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return REPORTS_DIR / f"report_{stamp}.xlsx"


def ensure_workbook(path: Path) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "results"
    ws.append(list(EXCEL_HEADERS))
    wb.save(path)


def append_result(path: Path, result: CaseResult) -> None:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    assert ws is not None
    ws.append(result.to_excel_row())
    wb.save(path)
